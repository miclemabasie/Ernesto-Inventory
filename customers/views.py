from django.shortcuts import render, redirect, get_object_or_404
from django.urls import reverse
from .models import Customer
from django.http import JsonResponse, HttpResponse
import json
from .forms import CustomerAddForm, SendMailForm
from django.contrib.auth.decorators import login_required
from openpyxl import Workbook
from datetime import datetime
import pywhatkit
from .tasks import test_func, send_whatsapp_mail_to_customer
from datetime import datetime


@login_required
def customer_list_view(request):
    customers = Customer.objects.all()
    form = CustomerAddForm()
    template_name = "customers/list.html"
    context = {
        "customers": customers,
        "customer_add_form": form,
    }

    return render(request, template_name, context)


@login_required
def customer_add_view(request):
    if request.method == "POST":
        data = json.loads(request.body)
        print("######################", data)
        # Create a new customer
        name = data["name"]
        phone = data["phone"]
        address = "None"
        email = "None"

        # Check if customer exists in database with same credentials (name, and phone)
        queryset = Customer.objects.filter(name=name, phone=phone)
        if len(queryset) > 0:
            data = {"status": "warning", "message": "Customer already exists!"}
            return JsonResponse(data)

        # Verify if address and email was provided in the fronend since there are optional
        if data["address"]:
            address = data["address"]
        if data["email"]:
            email = data["email"]

        customer = Customer.objects.create(
            name=name, phone=phone, address=address, email=email
        )
        customer.save()
        data = {"status": "success", "message": "Customer created successfully!"}
        return JsonResponse(data)
    return redirect("customers:list")


@login_required
def edit_customer_view(request, customer_id):
    """
    Edit customer information
    """
    customer = get_object_or_404(Customer, id=customer_id)
    if request.method == "POST":
        form = CustomerAddForm(request.POST, instance=customer)
        if form.is_valid():
            # Save customer information and redirect user to customer list page
            form.save()
            return redirect("customers:list")
    # if the request is a GET, send the user to the edit page
    # prepolating the existing user information in the form form(instance=customer)
    form = CustomerAddForm(instance=customer)
    template_name = "customers/edit.html"
    context = {
        "customer": customer,
        "customer_add_form": form,
    }
    return render(request, template_name, context)


def backup_customer_data(request):
    """
    Backup all the data related to the customer
    """
    wb = Workbook()
    # grab the active worksheet
    ws = wb.active
    # Write the column headers
    ws.cell(row=1, column=1).value = "id"
    ws.cell(row=1, column=2).value = "Customer Name"
    ws.cell(row=1, column=3).value = "Phone"
    ws.cell(row=1, column=4).value = "address"
    ws.cell(row=1, column=5).value = "email"

    customers = Customer.objects.all()
    index = 2
    for customer in customers:
        ws.cell(row=index, column=1).value = str(customer.id)
        ws.cell(row=index, column=2).value = str(customer.name)
        ws.cell(row=index, column=3).value = str(customer.phone)
        ws.cell(row=index, column=4).value = str(customer.address)
        ws.cell(row=index, column=5).value = str(customer.email)
        index += 1
    # Save the spreadsheet and return it to the user
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    day = str(datetime.today().day)
    month = str(datetime.today().month)
    year = str(datetime.today().year)

    response["Content-Disposition"] = (
        "attachment; filename="
        + "customers"
        + "_report_"
        + day
        + "_"
        + month
        + "_"
        + year
        + ".xlsx"
    )
    wb.save(response)
    return response


def send_message(request, customer_id):
    """
    Send whatsapp customer to customer.
    """

    customer = get_object_or_404(Customer, id=customer_id)
    if request.method == "POST":
        # Get the customer's number
        message = request.POST.get("message")
        number = customer.phone
        number = f"+237{number}"
        if len(number) == 13:
            print("### Sending mail to ")
            try:
                # sending message to receiver
                # using pywhatkit
                send_whatsapp_mail_to_customer.delay(number, message)
                print("Successfully Sent!")
                return redirect(reverse("home"))

            except:
                print("An Unexpected Error!")
    template_name = "customers/send_mail.html"
    send_mail_form = SendMailForm()
    context = {
        "form": send_mail_form,
        "customer": customer,
    }

    return render(request, template_name, context)


def test(request):
    test_func.delay()
    return HttpResponse("Done")
