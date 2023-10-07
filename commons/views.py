from django.shortcuts import render, redirect
from django.urls import reverse
from openpyxl import Workbook, load_workbook
from django.http import HttpResponse
from datetime import datetime
from .backup_handlers import (
    export_product_data,
    export_category_date,
    export_customer_data,
    export_sale_data,
    export_salesitem_data,
)
from .import_handlers import (
    import_category_data,
    import_product_data,
    import_customer_data,
    import_sale_data,
    import_sale_item_data,
)
from django.contrib.auth.decorators import login_required


# Create your views here.
@login_required
def export_data(request):
    """
    Handle the backing up of data to a CSV file
    """
    wb = Workbook()
    # grab the active worksheet, Products, Categories, SalesItems, Sales, Customers
    product_sheet = wb.active
    product_sheet.title = "Products"
    category_sheet = wb.create_sheet("Categories")
    saleitems_sheet = wb.create_sheet("SaleItems")
    sale_sheet = wb.create_sheet("Sales")
    customers_sheet = wb.create_sheet("Customers")
    # Write the column headers
    export_product_data(product_sheet)
    export_category_date(category_sheet)
    export_customer_data(customers_sheet)
    export_salesitem_data(saleitems_sheet)
    export_sale_data(sale_sheet)
    # Save the spreadsheet and return it to the user
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    day = str(datetime.today().day)
    month = str(datetime.today().month)
    year = str(datetime.today().year)

    response["Content-Disposition"] = (
        "attachment; filename="
        + "data_backup"
        + "_report_"
        + day
        + "_"
        + month
        + "_"
        + year
        + ".xlsx"
    )
    wb.save(response)
    return response


def import_data(request):
    if request.method == "POST":
        # Get the file from the request object
        file = request.FILES.get("data")
        wb = load_workbook(filename=file)
        product_sheet = wb.active
        category_sheet = wb.get_sheet_by_name("Categories")
        saleitems_sheet = wb.get_sheet_by_name("SaleItems")
        sale_sheet = wb.get_sheet_by_name("Sales")
        customers_sheet = wb.get_sheet_by_name("Customers")

        print(
            product_sheet, category_sheet, saleitems_sheet, sale_sheet, customers_sheet
        )
        import_category_data(category_sheet)
        import_product_data(product_sheet, request)
        import_customer_data(customers_sheet)
        import_sale_data(sale_sheet, request)
        import_sale_item_data(saleitems_sheet)

        return redirect(reverse("home"))

    template_name = "core/import_data.html"
    context = {
        "name": "Data Import",
    }
    return render(request, template_name, context)
