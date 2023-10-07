from openpyxl import Workbook, load_workbook
from products.models import Product, Category
from customers.models import Customer
from sales.models import SaleItem, Sale


def import_category_data(ws):
    for row in ws.iter_rows(min_row=2, values_only=True):
        # extract data from row values
        name, slug, created, updated = row[1], row[2], row[3], row[4]

        try:
            category = Category.objects.get(name=name)
        except Category.DoesNotExist:
            # Create new
            category = Category.objects.create(
                name=name, slug=slug, created=created, updated=updated
            )
            category.save()


def import_product_data(ws, request):
    for row in ws.iter_rows(min_row=2, values_only=True):
        # extract data from row values
        name, user, category, description = row[1], row[2], row[3], row[4]
        slug, quantity, reorderlevel, active = row[5], row[6], row[7], row[8]
        created, updated = row[9], row[10]
        # Create product

        user = request.user
        #
        category = Category.objects.get(name=category)

        try:
            product = Product.objects.get(
                name__iexact=name, category__name__iexact=category
            )
        except Product.DoesNotExist:
            # Create product if not in database
            product = Product.objects.create(
                name=name,
                user=user,
                category=category,
                description=description,
                slug=slug,
                quantity=quantity,
                reorder_level=reorderlevel,
                active=active,
                created=created,
                updated=updated,
            )
            product.save()


def import_customer_data(ws):
    for row in ws.iter_rows(min_row=2, values_only=True):
        # extract data from row values
        name, phone, address, email = row[1], row[2], row[3], row[4]
        try:
            customer = Customer.objects.get(name__iexact=name, phone__iexact=phone)
        except Customer.DoesNotExist:
            # Create customer if not in database
            customer = Customer.objects.create(
                name=name, phone=phone, address=address, email=email
            )
            customer.save()


def import_sale_data(ws, request):
    for row in ws.iter_rows(min_row=2, values_only=True):
        # extract data from row values
        transaction_id, sales_man, total_sale_price, date_created = (
            row[1],
            row[2],
            row[3],
            row[4],
        )
        validated = row[5]

        try:
            sale = Sale.objects.get(transaction_id=transaction_id)
        except Sale.DoesNotExist:
            # Create sale object
            sale = Sale.objects.create(
                transaction_id=transaction_id,
                sales_man=request.user,
                total_sale_price=float(total_sale_price),
                date_created=date_created,
                validated=validated,
            )
            sale.save()


def import_sale_item_data(ws):
    for row in ws.iter_rows(min_row=2, values_only=True):
        # extract data from row values
        product, unit_price, total_price, sale_transaction_id, created, quantity = (
            row[1],
            row[2],
            row[3],
            row[4],
            row[5],
            row[6],
        )
        # Get the Sale object
        sale = Sale.objects.get(transaction_id=sale_transaction_id)
        product = Product.objects.get(name=product)
        print(sale, total_price, "#################################")
        try:
            saleitem = SaleItem.objects.get(
                product=product, sale__transaction_id=sale_transaction_id
            )
        except SaleItem.DoesNotExist:
            # Create new saleitem
            saleitem = SaleItem.objects.create(
                product=product,
                quantity=quantity,
                unit_price=float(unit_price),
                total_price=float(total_price),
                sale=sale,
                created=created,
            )
            saleitem.save()
