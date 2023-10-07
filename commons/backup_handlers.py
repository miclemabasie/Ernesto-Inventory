from openpyxl import Workbook, load_workbook
from products.models import Product, Category
from customers.models import Customer
from sales.models import SaleItem, Sale


def export_product_data(ws):
    ws.cell(row=1, column=1).value = "ID"
    ws.cell(row=1, column=2).value = "User"
    ws.cell(row=1, column=3).value = "Category"
    ws.cell(row=1, column=4).value = "Description"
    ws.cell(row=1, column=5).value = "Slug"
    ws.cell(row=1, column=6).value = "Quantity"
    ws.cell(row=1, column=7).value = "ReorderLevel"
    ws.cell(row=1, column=8).value = "Active"
    ws.cell(row=1, column=9).value = "Created"
    ws.cell(row=1, column=10).value = "Updated"

    products = Product.objects.all()
    # get the lenght of the data
    data_len = len(products)
    index = 2
    for product in products:
        ws.cell(row=index, column=1).value = str(product.id)
        ws.cell(row=index, column=2).value = str(product.user.username)
        ws.cell(row=index, column=3).value = str(product.category)
        ws.cell(row=index, column=4).value = str(product.description)
        ws.cell(row=index, column=5).value = str(product.slug)
        ws.cell(row=index, column=6).value = str(product.quantity)
        ws.cell(row=index, column=7).value = str(product.reorder_level)
        ws.cell(row=index, column=8).value = str(product.active)
        ws.cell(row=index, column=9).value = str(product.created)
        ws.cell(row=index, column=10).value = str(product.updated)
        index += 1


def export_category_date(ws):
    ws.cell(row=1, column=1).value = "ID"
    ws.cell(row=1, column=2).value = "Name"
    ws.cell(row=1, column=3).value = "Slug"
    ws.cell(row=1, column=4).value = "Created"
    ws.cell(row=1, column=5).value = "Updated"

    categories = Category.objects.all()
    # get the lenght of the data
    index = 2
    for category in categories:
        ws.cell(row=index, column=1).value = str(category.id)
        ws.cell(row=index, column=2).value = str(category.name)
        ws.cell(row=index, column=3).value = str(category.slug)
        ws.cell(row=index, column=4).value = str(category.created)
        ws.cell(row=index, column=5).value = str(category.updated)
        index += 1


def export_customer_data(ws):
    """
    Backup all the data related to the customer
    """
    ws.cell(row=1, column=1).value = "ID"
    ws.cell(row=1, column=2).value = "Customer Name"
    ws.cell(row=1, column=3).value = "Phone"
    ws.cell(row=1, column=4).value = "address"
    ws.cell(row=1, column=5).value = "email"

    customers = Customer.objects.all()
    index = 2
    for customer in customers:
        ws.cell(row=index, column=1).value = str(customer.id)
        ws.cell(row=index, column=2).value = str(customer.name)
        ws.cell(row=index, column=3).value = str(customer.phone)
        ws.cell(row=index, column=4).value = str(customer.address)
        ws.cell(row=index, column=5).value = str(customer.email)
        index += 1


def export_salesitem_data(ws):
    """
    Backup data about saletimes
    """
    ws.cell(row=1, column=1).value = "id"
    ws.cell(row=1, column=2).value = "Product Name"
    ws.cell(row=1, column=3).value = "Unit Price"
    ws.cell(row=1, column=4).value = "Total Price"
    ws.cell(row=1, column=5).value = "Sale Transaction ID"
    ws.cell(row=1, column=6).value = "Date Created"

    sales = SaleItem.objects.filter(sale__validated=True)
    print("######", sales[0].product)
    index = 2
    for sale in sales:
        ws.cell(row=index, column=1).value = str(sale.id)
        ws.cell(row=index, column=2).value = str(sale.product.name)
        ws.cell(row=index, column=3).value = str(sale.unit_price)
        ws.cell(row=index, column=4).value = str(sale.total_price)
        ws.cell(row=index, column=5).value = str(sale.sale.transaction_id)
        ws.cell(row=index, column=6).value = str(sale.created)
        index += 1


def export_sale_data(ws):
    """
    Backup data about sale objects
    """
    ws.cell(row=1, column=1).value = "ID"
    ws.cell(row=1, column=2).value = "Transaction ID"
    ws.cell(row=1, column=3).value = "Sales Man"
    ws.cell(row=1, column=4).value = "Total Sale Price"
    ws.cell(row=1, column=5).value = "Date"
    ws.cell(row=1, column=6).value = "Validated"

    sales = Sale.objects.filter(validated=True)
    print(sales)
    index = 2
    for sale in sales:
        ws.cell(row=index, column=1).value = str(sale.id)
        ws.cell(row=index, column=2).value = str(sale.transaction_id)
        ws.cell(row=index, column=3).value = str(sale.sales_man)
        ws.cell(row=index, column=4).value = str(sale.total_sale_price)
        ws.cell(row=index, column=5).value = str(sale.date_created)
        ws.cell(row=index, column=6).value = str(sale.validated)
        index += 1
