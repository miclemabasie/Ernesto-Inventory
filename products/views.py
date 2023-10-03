from django.shortcuts import render, get_object_or_404, redirect
from .models import Product, Category
from django.contrib.auth.decorators import login_required
from django.http import JsonResponse, HttpResponse
from .decorators import require_post
from .forms import ProductAddForm, CategoryAddForm
import json
from django.utils.text import slugify
from django.urls import reverse
from django.core.exceptions import ValidationError
from sales.models import SaleItem
from django.db.models import Sum
from .utils import get_sales_data
from django.views.decorators.csrf import csrf_exempt
import json
from openpyxl import Workbook
from datetime import datetime


@csrf_exempt
@login_required
def home_view(request):
    template_name = "core/home.html"
    period = "month"
    sales_data = get_sales_data(period)
    if request.method == "POST":
        data = {}
        reqData = json.loads(request.body)
        print(reqData)

        chartType = reqData["chartType"]
        data.update({"sales_data": sales_data})
        data["chartType"] = "bar"
        data["values"] = sales_data["values"]
        data["labels"] = sales_data["labels"]
        return JsonResponse(data)
    context = {
        "section": "home",
        "values": sales_data["values"],
        "labels": sales_data["labels"],
    }
    return render(request, template_name, context)


@login_required
def product_list(request):
    template_name = "products/list.html"
    products = Product.objects.all()
    product_add_form = ProductAddForm()
    category_add_form = CategoryAddForm()
    categories = Category.objects.all()
    context = {
        "section": "Products",
        "products": products,
        "categories": categories,
        "product_form": product_add_form,
        "category_form": category_add_form,
    }

    return render(request, template_name, context)


@login_required
@require_post
def products_add_view(request):
    # Get the data sent from javascript
    data = json.loads(request.body)
    name = data["name"]
    category = data["category"]
    price = data["price"]
    quantity = data["quantity"]
    reorder_level = data["reorder_level"]
    # Try to get the category object of using the given category id
    try:
        category = Category.objects.get(id=category)
    except Category.DoesNotExist:
        return JsonResponse({"message": "invalid Category", "status": "error"})
    # Create a new product with data obj
    product = Product(
        name=name,
        category=category,
        price=price,
        quantity=quantity,
        reorder_level=reorder_level,
    )

    # Check if product already exists with same name
    try:
        test_product = Product.objects.get(name__iexact=name)
        if test_product:
            return JsonResponse(
                {"message": "Product with this name already exists", "status": "error"}
            )
    except Product.DoesNotExist:
        pass

    # Set the user to the current logged in user
    product.user = request.user
    # Set active to the product if quantity is greater than 0
    if int(product.quantity) < 0:
        product.active = False
    product.save()
    data = {
        "status": "success",
        "message": f"Succes: {product.name} added to {product.category.name}!",
    }
    return JsonResponse(data)


@login_required
def product_detail_view(request, product_id):
    product = get_object_or_404(Product, id=product_id)
    template_name = "products/detail.html"
    context = {
        "product": product,
    }
    return render(request, template_name, context)


@login_required
def update_product_view(request, product_id):
    product = get_object_or_404(Product, id=product_id)
    if request.method == "POST":
        form = ProductAddForm(request.POST, instance=product)
        if form.is_valid():
            form.save()
            return redirect(f"/products/?updatedproduct={product_id}")
            # return redirect(reverse("products:list"))
        else:
            raise ValidationError("Invalid product data")
    form = ProductAddForm(instance=product)
    template_name = "products/update.html"
    context = {
        "product": "product",
        "product_form": form,
    }
    return render(request, template_name, context)


@login_required
def delete_product_view(request, product_id):
    product = get_object_or_404(Product, id=product_id)
    # reqeust to delete product
    product.delete()
    return redirect(reverse("products:list"))


@login_required
@require_post
def category_add_view(request):
    print("Request to add")
    data = json.loads(request.body)
    category_name = data["name"]

    # Check if category with same name alread exists
    try:
        category = Category.objects.get(name__iexact=category_name)
        return JsonResponse(
            {"status": "Error", "message": "Category with this name already exist!"}
        )
    except Category.DoesNotExist:
        pass
    # Create the category
    category = Category(name=category_name)
    category.slug = slugify(category_name)
    category.save()
    data = {
        "status": "success",
        "message": "Category added successfuly",
    }
    return JsonResponse(data)


@login_required
def category_edit_view(request, category_id):
    category = get_object_or_404(Category, id=category_id)
    if request.method == "POST":
        form = CategoryAddForm(request.POST, instance=category)
        if form.is_valid():
            form.save()
            return redirect(reverse("products:list"))
    form = CategoryAddForm(instance=category)
    template_name = "products/category_edit.html"
    context = {
        "category_form": form,
    }
    return render(request, template_name, context)


@login_required
def delete_category_view(request, category_id):
    cateogry = get_object_or_404(Category, id=category_id)
    # reqeust to delete product
    cateogry.delete()
    return redirect(reverse("products:list"))


def export_product_data(request):
    """
    Handle the backing up of data to a CSV file
    """
    wb = Workbook()
    # grab the active worksheet
    ws = wb.active
    # Write the column headers
    ws.cell(row=1, column=1).value = "id"
    ws.cell(row=1, column=2).value = "User"
    ws.cell(row=1, column=3).value = "Category"
    ws.cell(row=1, column=4).value = "Description"
    ws.cell(row=1, column=5).value = "Slug"
    ws.cell(row=1, column=6).value = "Quantity"
    ws.cell(row=1, column=7).value = "ReorderLevel"
    ws.cell(row=1, column=8).value = "Active"
    ws.cell(row=1, column=9).value = "Created"
    ws.cell(row=1, column=10).value = "Updated"

    products = Product.objects.all()
    # get the lenght of the data
    data_len = len(products)
    index = 2
    for product in products:
        ws.cell(row=index, column=1).value = str(product.id)
        ws.cell(row=index, column=2).value = str(product.user.username)
        ws.cell(row=index, column=3).value = str(product.category)
        ws.cell(row=index, column=4).value = str(product.description)
        ws.cell(row=index, column=5).value = str(product.slug)
        ws.cell(row=index, column=6).value = str(product.quantity)
        ws.cell(row=index, column=7).value = str(product.reorder_level)
        ws.cell(row=index, column=8).value = str(product.active)
        ws.cell(row=index, column=9).value = str(product.created)
        ws.cell(row=index, column=10).value = str(product.updated)
        index += 1
    # Save the spreadsheet and return it to the user
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    day = str(datetime.today().day)
    month = str(datetime.today().month)
    year = str(datetime.today().year)

    response["Content-Disposition"] = (
        "attachment; filename="
        + "products"
        + "_report_"
        + day
        + "_"
        + month
        + "_"
        + year
        + ".xlsx"
    )
    wb.save(response)
    return response
